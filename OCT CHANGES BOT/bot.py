import os
import time
from datetime import datetime, timedelta
import psutil
import pyautogui
import subprocess
import win32com.client
from dmsService import Dms
from tulipService import Bot
from dataclasses import dataclass
from tulipService.model.variableModel import VariableModel
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Alignment


@dataclass
class BotInputSchema:
    # Key: VariableModel = None
    ...
    dmscred: VariableModel = None
    sapCred: VariableModel = None
    language: VariableModel = None
    server_name: VariableModel = None
    client: VariableModel = None
    beekeeperUrl: VariableModel = None
    transaction_code: VariableModel = None
    companycode: VariableModel = None
    changed_on: VariableModel = None
    Company_code_Screenshot: VariableModel = None
    Report: VariableModel = None
    excel_file: VariableModel = None
    start_date: VariableModel = None
    end_date: VariableModel = None
    reportEndImage: VariableModel = None
    creditLimitEntriesReport: VariableModel = None


@dataclass
class BotOutputSchema:
    def __init__(self):
        # self.Key = VariableModel()
        self.changed_on = VariableModel()
        self.Company_code_Screenshot = VariableModel()
        self.Report = VariableModel()
        # self.ExcelReport = VariableModel()
        self.otc28DonwloadedReport = VariableModel()
        self.combinedRowCount = VariableModel()
        self.creditLimitEntriesReport = VariableModel()
        self.creditLimitRowCount = VariableModel()
        self.riskCategoryRowCount = VariableModel()
        self.totalRowCount = VariableModel()
        self.start_date = VariableModel()
        self.end_date = VariableModel()
        self.reportEndImage=VariableModel()
        self.dataFound= VariableModel()


class SAPGUI:
    def __init__(self, server, client, user, password, lang, logger):
        self.log = logger
        self.sap_gui_app = None
        self.application = None
        self.connection = None
        self.launch_sap_gui()
        self.session = self.connect_to_sap(server_name=server, client=client,
                                           username=user, password=password, language=lang)

    def is_process_running(self, name: str, retry_count=0, delay=0) -> bool:
        """
        Check if a process with the given name is running with retry logic.

        :param name: The name of the process to check.
        :param retry_count: Number of retry attempts.
        :param delay: Delay between retry attempts in seconds.
        :return: True if the process is running, False otherwise.
        """
        for attempt in range(retry_count + 1):
            try:
                with subprocess.Popen(['tasklist', '/NH', '/FO', 'CSV'], stdout=subprocess.PIPE,
                                      encoding='utf-8') as tasklist:
                    for line in tasklist.stdout:
                        if name.lower() in line.lower():
                            self.log.info("Process found.")
                            return True
            except Exception as error:
                self.log.error(f"An error occurred: {error}")

            if attempt < retry_count:
                self.log.info("Retrying...")
                time.sleep(delay)

        return False

    def terminate_process(self, name: str):
        """
        Terminates the process with the given name.

        :param name: The name of the process to terminate.
        """
        try:
            for proc in psutil.process_iter(['pid', 'name']):
                if proc.info['name'].lower() == name.lower():
                    self.log.info(f"Terminating process {name} with PID {proc.info['pid']}.")
                    proc.terminate()
                    proc.wait(timeout=5)  # Wait for process to terminate
                    self.log.info(f"Process {name} terminated.")
                    return True
            self.log.info(f"No process with the name {name} found to terminate.")
            return False
        except Exception as error:
            self.log.error(
                f"An error occurred while terminating the process: {error}")
            return False

    # launching sapgui
    def launch_sap_gui(self):
        process_name = "saplogon.exe"
        self.log.info("start check process")
        if self.is_process_running(process_name, retry_count=3, delay=2):
            self.terminate_process(process_name)
            self.log.info("killing check process")
        self.log.info("after killing check process")
        os.startfile(process_name)
        # Wait for the SAP Logon process to be fully open and the SAPGUI object to be available
        retry_attempts = 8  # Adjust the number of retry attempts as needed
        sap_gui_available = False

        while retry_attempts > 0:
            if self.is_process_running(process_name, retry_count=1, delay=1):
                try:
                    sap_gui = win32com.client.GetObject("SAPGUI")
                    if sap_gui is not None:
                        self.log.info("SAP GUI is available.")
                        sap_gui_available = True
                        break
                except Exception as e:
                    self.log.info("Waiting for SAP GUI to become available...")
            else:
                self.log.info("Waiting for SAP Logon to open...")

            time.sleep(2)
            retry_attempts -= 1

        if not sap_gui_available:
            self.log.error("SAP GUI did not become available within the expected time.")

    # connect and login to sap
    def connect_to_sap(self, server_name, client, username, password, language):
        try:
            self.sap_gui_app = win32com.client.GetObject("SAPGUI")
            if not isinstance(self.sap_gui_app, win32com.client.CDispatch):
                return None

            application = self.sap_gui_app.GetScriptingEngine
            self.connection = application.OpenConnection(server_name, True)
            time.sleep(5)
            self.session = self.connection.Children(0)
            self.session.findById("wnd[0]").maximize()
            self.session.findById(
                "wnd[0]/usr/txtRSYST-BNAME").text = username
            self.session.findById(
                "wnd[0]/usr/pwdRSYST-BCODE").text = password
            self.session.findById(
                "wnd[0]/usr/txtRSYST-MANDT").text = client
            self.session.findById(
                "wnd[0]/usr/txtRSYST-LANGU").text = language
            self.session.findById("wnd[0]").sendVKey(0)
            return self.session
        except Exception as error:
            self.log.error(f"Error connecting to SAP: {error}")
            return None

    def disconnect(self):
        try:
            if self.connection:
                self.session.CloseSession()
        except Exception as error:
            self.log.error("Error disconnecting from SAP: {error}")
            raise error


class BotLogic(Bot):
    def __init__(self) -> None:
        super().__init__()
        try:
            # Initialize an instance of BotOutputSchema
            self.outputs = BotOutputSchema()

            # Fetch the proposed bot inputs based on the schema
            self.input = self.bot_input.get_proposedBotInputs(BotInputs=BotInputSchema)

            self._dms_identity = self.bot_input.get_identity(self.input.dmscred.value)
            self.sapidentity = self.bot_input.get_identity(self.input.sapCred.value)
            self.sap_gui = SAPGUI(server=self.input.server_name.value, client=self.input.client.value,
                                  user=self.sapidentity.credential.basicAuth.username,
                                  password=self.sapidentity.credential.basicAuth.password,
                                  lang=self.input.language.value,
                                  logger=self.log)

            self._dms = Dms(beekeeper_url=self.input.beekeeperUrl.value,
                            user_name=self._dms_identity.credential.basicAuth.username,
                            password=self._dms_identity.credential.basicAuth.password, logger=self.log)

        except Exception as error:
            # Log the error
            self.log.error(f"Error initializing BotLogic: {error}")

    def dms_upload_screenshot(self, screenshot_path):
        """this function returns the signature of the file
        """
        try:
            self.log.info(f"dms screenshot path inside the functions {screenshot_path}")
            screenshot_fs = self._dms.upload_file_to_dms(complete_path=screenshot_path)

            self.log.info(f"dms screenshot path{screenshot_fs}")
            self.log.info("screenshot is uploaded in DMS")
            # self._dms.download_file_dms()
            return screenshot_fs
        except Exception as error:
            self.bot_output.error(error)
            raise f"Error occured during dms upload: {error}"

    def take_and_upload_screenshot(self, screenshot_name):
        """

        :param screenshot_name:
        :return: it passes the Scrennshot file into Dms
        """
        try:
            screenshot_filename = f"{screenshot_name}.png"
            full_screenshot_path = './'
            screenshot_path = os.path.join(full_screenshot_path, screenshot_filename)
            self.log.info(f"screenshot name: {screenshot_filename}")
            time.sleep(3)
            screenshot = pyautogui.screenshot()
            screenshot.save(screenshot_path)
            self.log.info(f"Screenshot saved successfully at {screenshot_path}.")
            screenshot_file = self.dms_upload_screenshot(screenshot_path=screenshot_path)

            self.log.info(f"Screenshot uploaded successfully: {screenshot_file}")
            return screenshot_file
        except Exception as error:
            self.log.error(f"Error taking and uploading screenshot: {error}")
            raise f"Error taking and uploading screenshot: {error}"

    def company_code_section(self):
        """this functions has hanlde the multiple code selection"""
        try:
            self.sap_gui.session.findById("wnd[0]/usr/btn%_KKBER_%_APP_%-VALU_PUSH").press()
            # self.sap_gui.session.findById("wnd[1]/tbar[0]/btn[16]").press()
            # Values to set for each text field
            company_code = self.input.companycode.value
            element_id = (
                "wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,0]"
            )

            # Set focus on the current element
            self.sap_gui.session.findById(element_id).setFocus()

            # Set the text field value for the current element
            self.sap_gui.session.findById(element_id).text = company_code[0]

            self.log.info(f"company_code type = {type(company_code)}")
            for i in range(1, len(company_code)):
                # Create the dynamic string for the element ID
                element_id = (
                    "wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE/ctxtRSCSEL_255-SLOW_I[1,1]")

                # Set focus on the current element
                self.sap_gui.session.findById(element_id).setFocus()

                # Set the text field value for the current element
                self.sap_gui.session.findById(element_id).text = company_code[i]

                element = self.sap_gui.session.findById(
                    "wnd[1]/usr/tabsTAB_STRIP/tabpSIVA/ssubSCREEN_HEADER:SAPLALDB:3010/tblSAPLALDBSINGLE"
                )

                # Set the scrollbar position dynamically based on the current index
                if i < len(company_code) - 1:
                    element.verticalScrollbar.Position = i + 1

        # except Exception as e:
        #     self.log.error(f"Error in company code: {e}")
        #     self.bot_output.error(e)
        except Exception as e:
            self.log.info(f"Error occured in the company code selection functions: {e}")
        # self.bot_output.error(e)

    def convert_date_format(self, date_str):
        """this functions returns required format of the date's """
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            return date_obj.strftime('%d.%m.%Y')
        except ValueError as e:
            self.log.error(f"Error in date format conversion: {e}")
            return None
    def convert_date_format_invert(self, date_str):
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            return date_obj.strftime('%Y-%m-%d')
        except ValueError as e:
            self.log.error(f"Error in date format conversion: {e}")
            return None

    def get_previous_month_date_range(self):
        """this functions returns the start date, end date of previous month"""
        # Get today's date
        today = datetime.today()

        # Calculate the first day of the current month
        first_day_of_current_month = today.replace(day=1)

        # Calculate the last day of the previous month
        last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)

        # Calculate the first day of the previous month
        first_day_of_previous_month = last_day_of_previous_month.replace(day=1)

        # Format the dates as required
        start_date = first_day_of_previous_month.strftime('%d.%m.%Y')
        end_date = last_day_of_previous_month.strftime('%d.%m.%Y')

        return start_date, end_date

    def count_filtered_rows_and_export(self, file_path, sheet_name, column_letter, filter_value, output_file_path):

        # Load the workbook and select the sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Get the column index (1-based)
        column_index = column_index_from_string(column_letter)

        # Prepare a list to hold the filtered rows
        filtered_rows = []

        # Add the header row to the filtered_rows list
        header = [cell.value for cell in sheet[1]]
        filtered_rows.append(header)

        # Iterate through the rows in the sheet
        for row in sheet.iter_rows(values_only=True):
            if row[column_index - 1] == filter_value:
                filtered_rows.append(row)

        # Create a new workbook and add the filtered data
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "Filtered Data"

        # Write the filtered rows to the new sheet
        for row in filtered_rows:
            new_sheet.append(row)

        # Auto-adjust column widths
        for col in new_sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            new_sheet.column_dimensions[column].width = adjusted_width

        # Ensure that dates are correctly formatted
        for cell in new_sheet['A']:
            cell.number_format = 'DD/MM/YYYY'
            cell.alignment = Alignment(horizontal='left')

        # Save the new workbook
        new_workbook.save(output_file_path)

        # Return the count of filtered rows excluding the header

        return len(filtered_rows) - 1

    def giveDateRange(self):
        inputStartDate = self.input.start_date.value
        inputEndDate = self.input.end_date.value
        if(inputStartDate == "" or inputEndDate==""):
            startDate, endDate = self.get_previous_month_date_range()
        else:
            startDate = self.convert_date_format(inputStartDate)
            endDate = self.convert_date_format(inputEndDate)
        return startDate,endDate

    def main(self):

        """Bot Logic code"""
        try:
            self.sap_gui.session.findById("wnd[0]/tbar[0]/okcd").text = self.input.transaction_code.value

            self.sap_gui.session.findById("wnd[0]").sendVKey(0)

            self.company_code_section()
            company_upload = self.take_and_upload_screenshot(screenshot_name=self.input.Company_code_Screenshot.value)
            self.outputs.Company_code_Screenshot.value = company_upload

            self.sap_gui.session.findById("wnd[1]/tbar[0]/btn[8]").press()

            start,end = self.giveDateRange()
            self.sap_gui.session.findById("wnd[0]/usr/ctxtDATUM-LOW").text = start

            self.sap_gui.session.findById("wnd[0]/usr/ctxtDATUM-HIGH").text = end



            changed_upload = self.take_and_upload_screenshot(screenshot_name=self.input.changed_on.value)
            self.outputs.changed_on.value = changed_upload
            time.sleep(4)
            self.sap_gui.session.findById("wnd[0]/tbar[1]/btn[8]").press()
            report_upload = self.take_and_upload_screenshot(screenshot_name=self.input.Report.value)
            self.outputs.Report.value = report_upload

            self.sap_gui.session.findById("wnd[0]").sendVKey(24)

            report_upload_2 = self.take_and_upload_screenshot(screenshot_name=self.input.reportEndImage.value)
            self.outputs.reportEndImage.value = report_upload_2

            self.sap_gui.session.findById("wnd[0]/tbar[1]/btn[43]").press()

            self.sap_gui.session.findById("wnd[1]/tbar[0]/btn[0]").press()
            self.sap_gui.session.findById("wnd[1]/tbar[0]/btn[0]").press()
            full_excel_path = self.working_dir
            self.sap_gui.session.findById("wnd[1]/usr/ctxtDY_PATH").text = full_excel_path

            self.sap_gui.session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = self.input.excel_file.value

            self.sap_gui.session.findById("wnd[1]/tbar[0]/btn[0]").press()
            time.sleep(7)

            excel_filename = self.input.excel_file.value

            excel_path = os.path.join(full_excel_path, excel_filename)
            # excel_filename.save(screenshot_path)
            self.log.info(f"Screenshot saved successfully at {excel_path}.")
            excel_fs = self._dms.upload_file_to_dms(complete_path=excel_path)
            self.log.info(f"excel uploaded successfully: {excel_fs}")
            # excel_file_Save = self.excel_upload(self,excel_filename)
            self.outputs.otc28DonwloadedReport.value = excel_fs

            file_path = self.working_dir + "\otc28DonwloadedReport.xlsx"
            sheet_name = "Sheet1"
            column_letter = "E"
            filter_value = "Credit limit"
            output_file_path = self.input.creditLimitEntriesReport.value
            # print(f"{file_path}")

            creditLimitRowCount = self.count_filtered_rows_and_export(file_path, sheet_name, column_letter,
                                                                      filter_value,
                                                                      output_file_path)

            creditLimitEntriesReport_path = os.path.join(full_excel_path, output_file_path)
            # excel_filename.save(screenshot_path)
            self.log.info(f"Screenshot saved successfully at {creditLimitEntriesReport_path}.")
            excel_fs = self._dms.upload_file_to_dms(complete_path=creditLimitEntriesReport_path)
            self.log.info(f"excel uploaded successfully: {creditLimitEntriesReport_path}")
            # excel_file_Save = self.excel_upload(self,excel_filename)
            self.outputs.creditLimitEntriesReport.value = excel_fs

            file_path = self.working_dir + "\otc28DonwloadedReport.xlsx"

            sheet_name = "Sheet1"
            column_letter = "E"
            filter_value = "Risk category"
            output_file_path = "filtered_data1.xlsx"

            riskCategoryRowCount = self.count_filtered_rows_and_export(file_path, sheet_name, column_letter,
                                                                       filter_value,
                                                                       output_file_path)
            self.log.info(f"Number of filtered rows: {riskCategoryRowCount}")
            combinedRowCount = riskCategoryRowCount + creditLimitRowCount
            self.outputs.riskCategoryRowCount.value = riskCategoryRowCount
            self.outputs.creditLimitRowCount.value = creditLimitRowCount

            self.outputs.combinedRowCount.value = combinedRowCount
            self.log.info(f"{riskCategoryRowCount}")
            workbook = openpyxl.load_workbook(self.working_dir + "\otc28DonwloadedReport.xlsx")
            active_sheet = workbook.active

            # Get the row count
            row_count = active_sheet.max_row

            # print(f'The active sheet has {row_count} rows.')
            # self.log.info(f"start and end date will be{start},{end}")
            self.outputs.totalRowCount.value = row_count

            if row_count <= 1:
                self.outputs.dataFound.value = "False"
            else:
                self.outputs.dataFound.value = "True"

            startDate = self.convert_date_format_invert(start)
            self.outputs.start_date.value = startDate
            endDate = self.convert_date_format_invert(end)
            self.outputs.end_date.value = endDate

            self.bot_output.success()

            # self.sap_gui.session.findById("wnd[1]/tbar[0]/btn[11]").press()
        except Exception as e:
            # self.sap_gui.log.error(f"An error occurred during the process: {e}")
            self.log.error(f"Error in main execution: {e}")
            self.bot_output.error(e)
        finally:
            # if self.sap_gui:
            self.sap_gui.session.findById("wnd[0]/tbar[0]/okcd").text = "/nex"
            self.sap_gui.session.findById("wnd[0]").sendVKey(0)
            result = os.system("taskkill /im excel.exe /f")
            if result == 0:
                self.log.info("Excel closed successfully.")
            else:
                self.log.info("No Excel been opened.")

            self.log.info("its entered into the finally block")